"""
한국 로또 6/45 통계 기반 예측기

매주 실행하면 최신 회차까지 자동으로 가져와 캐시에 저장하고,
과거 데이터의 통계적 특성을 반영한 5가지 전략으로 번호를 추천합니다.

사용법:
  python lotto_predictor.py            (캐시가 있으면 바로 예측)
  python lotto_predictor.py --sync     (API에서 최신 데이터 동기화)

주의: 로또는 독립 시행이므로 실제 당첨 확률을 높이지 못합니다.
      과거 데이터의 분포를 따르는 '합리적인 번호 생성기'로 사용하세요.
"""

import json
import os
import random
import socket
import ssl
import time
import urllib.request
from collections import Counter, defaultdict
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

CACHE_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "lotto_cache.json")
RECENT_WINDOW = 50  # 최근 트렌드 분석 범위

# GitHub Pages에서 전체 데이터를 한 번에 받아옴 (동행복권 API 불필요)
GITHUB_ALL_URL = "https://smok95.github.io/lotto/results/all.json"

# Windows Python에서 SSL 인증서 검증 실패 우회
try:
    _ssl_ctx = ssl.create_default_context()
    # certifi가 있으면 사용
    import certifi
    _ssl_ctx.load_verify_locations(certifi.where())
except Exception:
    _ssl_ctx = ssl.create_default_context()
    _ssl_ctx.check_hostname = False
    _ssl_ctx.verify_mode = ssl.CERT_NONE


def load_cache():
    if os.path.exists(CACHE_FILE):
        with open(CACHE_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return []


def save_cache(data):
    with open(CACHE_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def import_from_excel(xlsx_path):
    """기존 lotto_data_collector.py로 만든 Excel을 읽어 캐시로 변환."""
    if not HAS_OPENPYXL:
        print("   ⚠️  openpyxl 미설치 — pip install openpyxl")
        return []
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path, read_only=True)
    ws = wb["당첨번호"] if "당첨번호" in wb.sheetnames else wb.active
    data = []
    rows = ws.iter_rows(min_row=2, values_only=True)
    for row in rows:
        if not row or row[0] is None:
            continue
        try:
            data.append({
                "round": int(row[0]),
                "date": str(row[1]),
                "numbers": sorted([int(row[i]) for i in range(2, 8)]),
                "bonus": int(row[8]),
            })
        except (ValueError, TypeError, IndexError):
            continue
    data.sort(key=lambda x: x["round"])
    save_cache(data)
    print(f"   ✅ Excel에서 {len(data)}회차 가져옴 → {CACHE_FILE}")
    return data


BALL_COLORS = [
    (range(1, 11), "FFC000"),
    (range(11, 21), "4472C4"),
    (range(21, 31), "FF4444"),
    (range(31, 41), "A5A5A5"),
    (range(41, 46), "70AD47"),
]


def ball_color(n):
    for rng, color in BALL_COLORS:
        if n in rng:
            return color
    return "FFFFFF"


def export_excel(data, filename=None):
    """전체 당첨번호와 통계를 Excel로 저장 (캐시 백업용)."""
    if not HAS_OPENPYXL:
        print("   ⚠️  openpyxl 미설치 — Excel 백업 건너뜀 (pip install openpyxl)")
        return None
    if filename is None:
        filename = os.path.join(
            os.path.dirname(CACHE_FILE),
            f"lotto_backup_{data[-1]['round']}회_{datetime.now().strftime('%Y%m%d')}.xlsx",
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "당첨번호"

    headers = ["회차", "추첨일", "번호1", "번호2", "번호3", "번호4", "번호5", "번호6", "보너스"]
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Border(*[Side(style="thin") for _ in range(4)])

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center")
        c.border = thin

    num_font = Font(bold=True, size=11)
    for row_idx, d in enumerate(data, 2):
        ws.cell(row=row_idx, column=1, value=d["round"]).alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=2, value=d["date"]).alignment = Alignment(horizontal="center")
        for i, num in enumerate(d["numbers"]):
            c = ws.cell(row=row_idx, column=3 + i, value=num)
            c.alignment = Alignment(horizontal="center")
            color = ball_color(num)
            c.fill = PatternFill("solid", fgColor=color)
            c.font = Font(bold=True, size=11, color="FFFFFF" if color == "4472C4" else "000000")
        bc = ws.cell(row=row_idx, column=9, value=d["bonus"])
        bc.alignment = Alignment(horizontal="center")
        color = ball_color(d["bonus"])
        bc.fill = PatternFill("solid", fgColor=color)
        bc.font = Font(bold=True, size=11, color="FFFFFF" if color == "4472C4" else "000000")
        for col in range(1, 10):
            ws.cell(row=row_idx, column=col).border = thin

    for c, w in zip("ABCDEFGHI", [8, 14, 8, 8, 8, 8, 8, 8, 8]):
        ws.column_dimensions[c].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:I{len(data) + 1}"

    # 통계 시트
    ws2 = wb.create_sheet("번호별통계")
    for col, h in enumerate(["번호", "출현횟수", "보너스출현"], 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center")

    main_count = Counter()
    bonus_count = Counter()
    for d in data:
        main_count.update(d["numbers"])
        bonus_count[d["bonus"]] += 1
    for n in range(1, 46):
        ws2.cell(row=n + 1, column=1, value=n).alignment = Alignment(horizontal="center")
        ws2.cell(row=n + 1, column=2, value=main_count[n]).alignment = Alignment(horizontal="center")
        ws2.cell(row=n + 1, column=3, value=bonus_count[n]).alignment = Alignment(horizontal="center")

    wb.save(filename)
    print(f"   💾 Excel 백업: {os.path.basename(filename)}")
    return filename


def sync_data():
    """GitHub에서 전체 로또 데이터를 한 번에 받아 캐시에 저장합니다."""
    print(f"📡 GitHub에서 전체 데이터 다운로드 중...")
    print(f"   소스: smok95/lotto (GitHub Pages)")

    req = urllib.request.Request(GITHUB_ALL_URL, headers={
        "User-Agent": "Mozilla/5.0 lotto-predictor",
    })
    try:
        with urllib.request.urlopen(req, timeout=30, context=_ssl_ctx) as resp:
            raw = resp.read().decode("utf-8")
        github_data = json.loads(raw)
    except Exception as e:
        print(f"   ❌ 다운로드 실패: {e}")
        cache = load_cache()
        if cache:
            print(f"      기존 캐시 {len(cache)}회차로 진행합니다.\n")
            return cache
        print(f"      캐시도 없습니다. 네트워크를 확인하세요.")
        return []

    # GitHub 데이터 → 내부 캐시 형식으로 변환
    cache = []
    for item in github_data:
        date_str = item.get("date", "")
        if "T" in date_str:
            date_str = date_str.split("T")[0]
        cache.append({
            "round": item["draw_no"],
            "date": date_str,
            "numbers": sorted(item["numbers"]),
            "bonus": item["bonus_no"],
        })

    cache.sort(key=lambda x: x["round"])
    save_cache(cache)
    export_excel(cache)
    print(f"   ✅ 총 {len(cache)}회차 저장 완료 (1회 ~ {cache[-1]['round']}회)\n")
    return cache


# ============================================================
# 2. 통계 분석
# ============================================================

def analyze(data):
    total = len(data)
    freq_all = Counter()
    freq_recent = Counter()
    pair_count = defaultdict(int)
    last_seen = {}  # 번호별 마지막 등장 회차
    sums = []
    odd_counts = []
    range_dist = Counter()  # 1~10, 11~20, ...

    for d in data:
        nums = d["numbers"]
        freq_all.update(nums)
        sums.append(sum(nums))
        odd_counts.append(sum(1 for n in nums if n % 2 == 1))
        for n in nums:
            last_seen[n] = d["round"]
            range_dist[(n - 1) // 10] += 1
        for i in range(6):
            for j in range(i + 1, 6):
                a, b = sorted((nums[i], nums[j]))
                pair_count[(a, b)] += 1

    recent = data[-RECENT_WINDOW:] if len(data) >= RECENT_WINDOW else data
    for d in recent:
        freq_recent.update(d["numbers"])

    latest_round = data[-1]["round"]
    gap = {n: latest_round - last_seen.get(n, 0) for n in range(1, 46)}

    return {
        "total": total,
        "latest_round": latest_round,
        "freq_all": freq_all,
        "freq_recent": freq_recent,
        "pair_count": pair_count,
        "gap": gap,
        "avg_sum": sum(sums) / len(sums),
        "sum_min": int(sum(sums) / len(sums) - 35),
        "sum_max": int(sum(sums) / len(sums) + 35),
        "odd_mode": Counter(odd_counts).most_common(1)[0][0],
        "range_dist": range_dist,
    }


# ============================================================
# 3. 예측 전략
# ============================================================

def weighted_sample(weights):
    """가중치 기반으로 중복 없이 6개 선택."""
    nums = list(range(1, 46))
    w = [max(weights.get(n, 0.0), 0.001) for n in nums]
    picked = []
    while len(picked) < 6:
        choice = random.choices(nums, weights=w, k=1)[0]
        if choice not in picked:
            picked.append(choice)
            idx = nums.index(choice)
            w[idx] = 0
    return sorted(picked)


def is_balanced(nums, stats):
    s = sum(nums)
    if not (stats["sum_min"] <= s <= stats["sum_max"]):
        return False
    odd = sum(1 for n in nums if n % 2 == 1)
    if odd in (0, 6):
        return False
    # 한 구간(10단위)에 5개 이상 몰리지 않게
    rng = Counter((n - 1) // 10 for n in nums)
    if max(rng.values()) >= 5:
        return False
    # 연속 4개 이상 배제
    consec = 1
    max_consec = 1
    for i in range(1, 6):
        if nums[i] == nums[i - 1] + 1:
            consec += 1
            max_consec = max(max_consec, consec)
        else:
            consec = 1
    if max_consec >= 4:
        return False
    return True


def strategy_frequency(stats):
    """전체 빈도 가중 (핫넘버 선호)"""
    return weighted_sample({n: c for n, c in stats["freq_all"].items()})


def strategy_recent(stats):
    """최근 트렌드 가중"""
    weights = {n: stats["freq_recent"].get(n, 0) + 1 for n in range(1, 46)}
    return weighted_sample(weights)


def strategy_cold(stats):
    """오래 안 나온 번호 반등형"""
    return weighted_sample({n: g + 1 for n, g in stats["gap"].items()})


def strategy_pair(stats):
    """가장 자주 함께 나온 쌍을 씨앗으로"""
    top_pairs = sorted(stats["pair_count"].items(), key=lambda x: -x[1])[:20]
    seed_pair = random.choice(top_pairs)[0]
    picked = list(seed_pair)
    # 이 쌍과 함께 자주 등장한 번호를 가중치로
    partners = Counter()
    for (a, b), c in stats["pair_count"].items():
        if a in picked and b not in picked:
            partners[b] += c
        elif b in picked and a not in picked:
            partners[a] += c
    weights = {n: partners.get(n, 1) for n in range(1, 46) if n not in picked}
    while len(picked) < 6:
        nums = list(weights.keys())
        w = list(weights.values())
        pick = random.choices(nums, weights=w, k=1)[0]
        picked.append(pick)
        weights.pop(pick)
    return sorted(picked)


def strategy_balanced(stats):
    """전체 빈도 기반이되, 역사적 분포 제약을 만족할 때까지 재추첨"""
    for _ in range(500):
        nums = weighted_sample({n: c for n, c in stats["freq_all"].items()})
        if is_balanced(nums, stats):
            return nums
    return nums  # 제약 만족 실패 시 마지막 결과


STRATEGIES = [
    ("전체빈도 가중", strategy_frequency),
    ("최근트렌드 가중", strategy_recent),
    ("콜드넘버 반등", strategy_cold),
    ("동반출현 쌍기반", strategy_pair),
    ("균형형(역사분포)", strategy_balanced),
]


# ============================================================
# 4. 출력
# ============================================================

def print_stats(stats):
    print("📊 통계 요약")
    print(f"   분석 대상: 1회 ~ {stats['latest_round']}회 ({stats['total']}회차)")
    print(f"   평균 합계: {stats['avg_sum']:.1f}  (권장 범위 {stats['sum_min']}~{stats['sum_max']})")
    print(f"   최빈 홀수개수: {stats['odd_mode']}개")

    top_hot = stats["freq_all"].most_common(6)
    print(f"   🔥 핫넘버 TOP6: {[n for n, _ in top_hot]}")

    cold = sorted(stats["gap"].items(), key=lambda x: -x[1])[:6]
    print(f"   🧊 오래된 번호 TOP6: {[n for n, _ in cold]} (각각 {[g for _, g in cold]}회차 미출현)")

    recent_hot = stats["freq_recent"].most_common(6)
    print(f"   📈 최근 {RECENT_WINDOW}회 핫: {[n for n, _ in recent_hot]}")
    print()


def format_numbers(nums):
    return "  ".join(f"{n:>2d}" for n in nums)


def print_predictions(stats, sets_per_strategy=1):
    print(f"🎯 {stats['latest_round'] + 1}회차 추천 번호")
    print("=" * 55)
    for name, fn in STRATEGIES:
        for _ in range(sets_per_strategy):
            nums = fn(stats)
            s = sum(nums)
            odd = sum(1 for n in nums if n % 2 == 1)
            print(f"  [{name:<16}] {format_numbers(nums)}   합:{s} 홀:{odd}")
    print("=" * 55)
    print("⚠️  로또는 독립 시행입니다. 재미로만 활용하세요.\n")


# ============================================================
# 5. 메인
# ============================================================

def main():
    import sys
    print("🎱 한국 로또 6/45 예측기\n")

    # CLI: python lotto_predictor.py import <xlsx경로>
    if len(sys.argv) >= 3 and sys.argv[1] == "import":
        import_from_excel(sys.argv[2])
        return

    # --offline 플래그 또는 캐시가 이미 있으면 네트워크 없이 실행
    offline = "--offline" in sys.argv

    cache = load_cache()
    if offline and cache:
        print(f"📂 오프라인 모드: 캐시에서 {len(cache)}회차 로드\n")
        data = cache
    elif cache and not offline:
        # 캐시가 있으면 기본적으로 오프라인으로 동작, 동기화하려면 --sync 사용
        if "--sync" in sys.argv:
            data = sync_data()
        else:
            print(f"📂 캐시에서 {len(cache)}회차 로드 (최신 동기화: python lotto_predictor.py --sync)\n")
            data = cache
    else:
        # 캐시 없으면 동기화 시도
        data = sync_data()

    if not data:
        print("데이터가 없습니다.")
        print("  방법 1: Colab에서 lotto_colab_collector.ipynb 실행 → lotto_cache.json 다운로드")
        print("  방법 2: python lotto_predictor.py --sync (직접 API 수집)")
        return
    stats = analyze(data)
    print_stats(stats)
    print_predictions(stats)


if __name__ == "__main__":
    main()
